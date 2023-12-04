import os
import re
import uuid

import requests
from bs4 import BeautifulSoup
from django.contrib.sites.shortcuts import get_current_site
from django.core.files.storage import default_storage
from django.core.mail import EmailMessage
from django.db.models import Q
from django.utils import timezone
from django.core.files.base import ContentFile
from django_filters.rest_framework import DjangoFilterBackend
from jdatetime import datetime as jdatetime
from openpyxl import load_workbook
from rest_framework import viewsets, generics, filters, views, parsers
from rest_framework.response import Response

from authentication.permissions import IsSuperuserOrHR
from job.models import Requirement
from utils import config
from .models import CandidateModel, EducationModel, PreferencesModel, ExperiencesModel, AppointmentModel,\
    SettingsModel, StatusModel
from .serializers import ExcelFileSerializer, CandidateSerializer, ScoreSerializer, CandidateUpdateSerializer, \
    AppointmentSerializer, SettingsSerializer, PDFScoreSerializer


def calculate_skill_score(candidate, educations, requirement, experiences_count):
    """
        Calculate the skill score of a candidate based on their educations, skills, about, languages and experiences.

        Args:
            candidate: Candidate object.
            educations: List of EducationModel objects.
            requirement: List of Requirement objects.
            experiences_count: Number of experiences.
            
        Returns: Total skill score.
    """
    total_score = experiences_count
    print(experiences_count)
    for req in requirement:
        en, fa = req.en_title, req.fa_title
        req_counted = False
        for edu in educations:
            if en.lower() in ''.join([edu.level, edu.major]).lower() or fa in ''.join([edu.level, edu.major]):
                if not req_counted:
                    print(fa)
                    total_score += req.score
                    req_counted = True
                    break

        for skill in (candidate.languages, candidate.skills, candidate.about):
            if skill:
                if en.lower() in ''.join(skill).lower() or fa in ''.join(skill):
                    if not req_counted:
                        print(fa)
                        total_score += req.score
                        break
    return total_score


def create_appointment(candidate, start_time, end_time):
    """
        Create an appointment for a candidate and email the interview start and end times.

        Args:
             candidate: Candidate object.
             start_time: Start time of the interview.
             end_time: End time of the interview.
    """
    appointment = AppointmentModel.objects.create(candidate_id=candidate.id,
                                                  interview_start_time=start_time,
                                                  interview_end_time=end_time)
    print(appointment.interview_start_time, appointment.interview_end_time)

    # EmailMessage(f'Interview Invitation - {candidate.job}',
    #              f'Hello dear {candidate.name} Scheduled Interview: from {appointment.interview_start_time} to {appointment.interview_end_time}',
    #              config.EMAIL_HOST_USER, [candidate.email]).send()


def schedule_interviews(candidate, interview_duration_hours, start_work, end_work, current_date):
    """
    Schedule interviews for qualified candidates based on settings.

    Args:
         candidate: Candidate object.
         interview_duration_hours: Duration of the interview in hours.
         start_work: Start working time.
         end_work: End working time.
         current_date: Current date.
    """

    last_appointment = AppointmentModel.objects.all().order_by(
        '-interview_end_time').first()
    if last_appointment:
        if last_appointment.interview_end_time >= last_appointment.interview_end_time.replace(hour=end_work, minute=0,
                                                                                              second=0, microsecond=0):
            start_time = last_appointment.interview_end_time.replace(hour=start_work, minute=0, second=0,
                                                                     microsecond=0) + timezone.timedelta(days=1)
            end_time = start_time + timezone.timedelta(minutes=interview_duration_hours)
            create_appointment(candidate, start_time, end_time)
        else:
            start_time = last_appointment.interview_end_time
            end_time = start_time + timezone.timedelta(minutes=interview_duration_hours)
            create_appointment(candidate, start_time, end_time)

    else:
        if current_date >= current_date.replace(hour=start_work, minute=0, second=0, microsecond=0):

            start_time = current_date.replace(hour=start_work, minute=0, second=0,
                                                microsecond=0) + timezone.timedelta(days=1)
            end_time = start_time + timezone.timedelta(minutes=interview_duration_hours)
            create_appointment(candidate, start_time, end_time)
        else:
            start_time = current_date.replace(hour=start_work, minute=0, second=0, microsecond=0)
            end_time = start_time + timezone.timedelta(minutes=interview_duration_hours)
            create_appointment(candidate, start_time, end_time)


class UploadExcelAPIView(generics.CreateAPIView):
    """
        API View for uploading Excel files with candidate information.

        Handles parsing the link in Excel file and creating candidate objects.

    """
    serializer_class = ExcelFileSerializer
    # permission_classes = [IsSuperuserOrHR]
    parser_classes = (parsers.MultiPartParser,)

    def save_education(self, soup, candidate_id, education_to_save):
        for education in soup.select(
                'div.card-header:-soup-contains("تحصیلی") + div.card-body div.list-group-item label.d-block'):
            education_text = education.get_text(strip=True).split('-')
            if education_text:
                new_education = EducationModel(
                    candidate_id=candidate_id,
                    level=education_text[0].strip(),
                    major=education_text[1].strip(),
                    university=education.find_next('div', class_='font-size-base').find('span').get_text(strip=True),
                    duration=re.sub(r'\s+', ' ', education.find_next('div', class_='font-size-base').find('span',
                                                                                                          class_='mr-3').text),
                )
            else:
                None
            education_to_save.append(new_education)

    def save_preferences(self, soup, candidate_id, preferences_to_save):
        """
            Save preference information from the parsed HTML soup.

                 soup: Parsed HTML soup.
                 candidate_id: ID of the candidate.
                 preferences_to_save: List to store PreferencesModel objects.
        """
        output_dict = {}
        for preference in soup.select(
                'div.card-header:-soup-contains("ترجیحات") + div.card-body div.list-group-item label.d-block'):
            next_sibling = preference.find_next_sibling('div',
                                                        class_='font-size-2xl vertical-align-middle color-grey-light-1')
            values = next_sibling.find_all('label', class_='font-size-base color-grey-dark-2 mh-1')
            key = re.sub(r'\s+', ' ', preference.get_text(strip=True).replace('\u200c', ' ').strip())
            value = [re.sub(r'\s+', ' ', val.get_text(strip=True).replace('\u200c', ' ').replace('\n', '')) for val in
                     values]
            output_dict[key] = value

        new_preference = PreferencesModel(
            candidate_id=candidate_id,
            proviences=output_dict.get('استان های مورد نظر برای کار:'),
            job_category=output_dict.get('دسته بندی شغلی و زمینه کاری:'),
            job_level=output_dict.get('سطح ارشدیت در زمینه فعالیت:'),
            contracts_type=output_dict.get('نوع قراردادهای قابل قبول:'),
            salary=output_dict.get('حقوق مورد نظر:'),
            benefits=output_dict.get('مزایای شغلی مورد علاقه:') if 'مزایای شغلی مورد علاقه:' in output_dict else None
        )
        preferences_to_save.append(new_preference)

    def save_experiences(self, soup, candidate_id, experiences_to_save):

        for experience in soup.select(
                'div.card-header:-soup-contains("سوابق شغلی") + div.card-body div.list-group-item'):
            title = experience.find('label').text
            company = experience.find('span').text
            start_at = re.sub(r'\s+', ' ', experience.find('span', class_='mr-3').find('b').text)
            end_at = re.sub(r'\s+', ' ', experience.find('span', class_='mr-3').find_all('b')[-1].text.strip())

            start_date = jdatetime.strptime(start_at, "%B %Y").date()
            if end_at == "حالا":
                end_date = jdatetime.now().date()
            else:
                end_date = jdatetime.strptime(end_at, "%B %Y").date()

            new_experience = ExperiencesModel(
                candidate_id=candidate_id,
                title=title or None,
                company=company or None,
                start_at=start_at or None,
                end_at=end_at or None,
                duration=(end_date - start_date).days
            )
            experiences_to_save.append(new_experience)

    def extract_languages(self, soup):
        language_element = soup.select('div.card-header:-soup-contains("زبان") + div.card-body div.list-group-item')
        return [re.sub(r'\s+', ' ', language.find('label').get_text()) for language in
                language_element] if language_element else None

    def extract_skills(self, soup):
        skill_element = soup.select(
            'div.card-header:-soup-contains("حرفه") + div.card-body div.font-size-2xl.vertical-align-middle.color-grey-light-1 label.font-size-base.color-grey-dark-2.mh-1')
        return [skill.get_text(strip=True) for skill in skill_element] if skill_element else None

    def extract_data(self, soup, label_text):
        label = soup.find('label', text=re.compile(label_text))
        return re.sub(r'\s+', ' ', label.find_next_sibling('span').get_text(strip=True)) if label else None

    def process_row(self, row, file_path, user_email, candidates_to_create, experiences_to_save, preferences_to_save,
                    education_to_save):
        hyperlink = row[6].hyperlink

        if hyperlink is not None:
            try:
                response = requests.get(hyperlink.target)
                soup = BeautifulSoup(response.text, 'html.parser')
                url = soup.find(lambda tag: tag.name == 'a' and 'دانلود' in tag.get_text(strip=True))['href']
                resume_response = requests.get(url)
                unique_id = uuid.uuid1()
                candidate_data = {
                    'id': unique_id,
                    'name': row[2].value,
                    'job': row[1].value,
                    'phone_number': row[3].value,
                    'email': user_email,
                    'request_date': row[5].value,
                    'resume': ContentFile(resume_response.content,
                                          name=f"{row[1].value}/{row[2].value}.pdf") if resume_response else None,
                    'job_status': self.extract_data(soup, 'وضعیت اشتغال'),
                    'last_company': self.extract_data(soup, 'شرکت'),
                    'education_level': self.extract_data(soup, 'تحصیلی'),
                    'province': self.extract_data(soup, 'استان'),
                    'location': self.extract_data(soup, 'آدرس محل'),
                    'marital': self.extract_data(soup, 'تاهل'),
                    'birthdate': self.extract_data(soup, 'سال تولد'),
                    'gender': self.extract_data(soup, 'جنسیت'),
                    'military_service_status': self.extract_data(soup, 'وضعیت خدمت'),
                    'about': soup.find('p', class_='u-textJustify').get_text() if soup.find('p',
                                                                                            class_='u-textJustify') else None,
                    'skills': self.extract_skills(soup),
                    'languages': self.extract_languages(soup),
                }
                candidates_to_create.append(CandidateModel(**candidate_data))
                self.save_experiences(soup, unique_id, experiences_to_save)
                self.save_preferences(soup, unique_id, preferences_to_save)
                self.save_education(soup, unique_id, education_to_save)

            except :
                os.remove(file_path)
                return Response({'success': False, 'status': 422, 'error': 'Cannot access URL'})

    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            return Response({'success': False, 'status': 400, 'error': 'No file uploaded'})

        file_name = default_storage.save(uploaded_file, uploaded_file)
        file_path = default_storage.path(file_name)
        candidates_to_create = []
        experiences_to_save = []
        preferences_to_save = []
        education_to_save = []
        new_user_count = 0
        similar_candidate = 0
        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            existing_emails = set(CandidateModel.objects.values_list('email', flat=True))

            for row in worksheet.iter_rows(min_col=1, min_row=2):
                user_email = row[4].value

                if user_email in existing_emails:
                    similar_candidate += 1
                    continue
                self.process_row(row, file_path, user_email, candidates_to_create, experiences_to_save,
                                 preferences_to_save, education_to_save)

                new_user_count += 1

        except Exception as e:
            return Response({'success': False, 'status': 500, 'error': f'Error processing file: {str(e)}'})
        finally:
            os.remove(file_path)
            CandidateModel.objects.bulk_create(candidates_to_create)
            ExperiencesModel.objects.bulk_create(experiences_to_save)
            PreferencesModel.objects.bulk_create(preferences_to_save)
            EducationModel.objects.bulk_create(education_to_save)
            message = {
                'message': 'Data uploaded',
                'count new users': new_user_count,
                'count candidates exist': similar_candidate
            }
            return Response({'success': True, 'status': 200, 'message': message})


class NewCandidateScoreAPIView(generics.ListAPIView):
    """
    Calculate and update the skill scores for candidates.

    Methods:
        get(self, request, *args, **kwargs):Scores and sends a rejection email if the candidate's score is zero.
    """
    queryset = CandidateModel.objects.prefetch_related('experiencesmodel_set', 'educationmodel_set').filter(score=None)
    serializer_class = ScoreSerializer
    # permission_classes = [IsSuperuserOrHR, IsSuperuserOrTD]

    def get(self, request, *args, **kwargs):
        candidates = self.get_queryset()
        requirement = Requirement.objects.all()

        count = 0
        updated_candidates = []
        for candidate in candidates:
            educations = candidate.educationmodel_set.all()
            experiences_count = candidate.experiencesmodel_set.count()
            candidate.score = calculate_skill_score(candidate, educations, requirement, experiences_count)
            updated_candidates.append(candidate)
            count += 1

        CandidateModel.objects.bulk_update(updated_candidates, ['score'])

        if count == 0:
            return Response(
                {'success': True, 'status': 400, 'error': 'All candidates have already scored.'})

        return Response({'success': True, 'status': 200, 'message': f'{count} candidates scored'})


class SchedulerAPIView(views.APIView):
    """
        Schedule interviews for new candidates.

        Methods:
            get(): Schedule interviews based on candidate scores .
    """
    def get(self, request):
        candidates = CandidateModel.objects.filter(appointmentmodel__interview_start_time__isnull=True)

        current_date = timezone.now()
        settings = SettingsModel.objects.all().first()

        count_appointment = 0
        no_score=0
        if settings is None:
            return Response({'success': False, 'status': 400, 'error': 'You have no data in settings.'})
        for candidate in candidates:
            if candidate.PDF_score is None:
                no_score+=1
                continue
            if candidate.PDF_score >= settings.pass_score:
                count_appointment += 1
                print('accepted:Send Interview Invitation date')

                schedule_interviews(candidate, settings.interview_duration_hours, settings.start_work_time,
                                    settings.end_work_time, current_date)
                try:
                    status_model = StatusModel.objects.get(candidate_id=candidate.id)
                    status_model.status = 'WI'
                    status_model.save()
                except StatusModel.DoesNotExist:
                    StatusModel.objects.create(candidate_id=candidate.id, status='WI')

            else:
                print('rejected')
                try:
                    status_model = StatusModel.objects.get(candidate_id=candidate.id)
                    status_model.status = 'R'
                    status_model.save()
                except StatusModel.DoesNotExist:
                    StatusModel.objects.create(candidate_id=candidate.id, status='R')

                # EmailMessage(f'Your resume rejected', 'Hello, we may reach out to you again in the future',
                #              config.EMAIL_HOST_USER, [candidate.email]).send()

        if count_appointment == 0:
            message = {
                'error': 'No interview scheduled!',
                'reason1': 'All eligible candidates have already scheduled',
                'reason2': 'HR have not scored to PDFs',
                'reason3': 'No eligible candidates in DB'
            }
            return Response(
                {'success': True, 'status': 400, 'error': message})
        message = {
            'message': f'Interview time scheduled for {count_appointment} candidates',
            'have no score': f'{no_score} candidates',
        }
        return Response({'success': True, 'status': 200, 'message': message})


class CandidateListAPIView(generics.ListAPIView):
    queryset = CandidateModel.objects.all()
    serializer_class = CandidateSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['job']
    ordering_fields = ['request_date']

    # permission_classes = [IsSuperuserOrHR]


class OldCandidateInvitationAPIView(generics.ListAPIView):
    """
        Send invitations to old candidates for profile updates.

        Methods:
            get(self, request, *args, **kwargs): Send invitations and list candidates.
     """
    queryset = CandidateModel.objects.filter(Q(appointmentmodel__interview_start_time__isnull=True) | Q(statusmodel__status='R'))
    serializer_class = CandidateSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ('job',)
    # permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        job_param = request.GET.get('job', '')
        candidates = CandidateModel.objects.filter(job=job_param).filter(Q(appointmentmodel__interview_start_time__isnull=True) | Q(statusmodel__status='R'))
        current_site = get_current_site(self.request)
        for candidate in candidates:
            self.send_invitation_email(candidate, job_param, current_site, candidate.id)

        return self.list(request, *args, **kwargs)

    def send_invitation_email(self, candidate, job_param, current_site, candidate_id):
        print(
            f'Hello {candidate.name} Please consider updating your resume and applying for {job_param} position if you are interested Click on the following link to apply:http://{current_site.domain}/candidate/candidate_update/{candidate_id}', )

        # EmailMessage('Job Opportunity Update',
        #              f'Hello {candidate.name} \n' f'Please consider updating your resume and applying for {job_param} position if you are interested.\n Click on the following link to apply:http://{current_site.domain}/candidate/candidate_update/{candidate_id}',
        #              config.EMAIL_HOST_USER, [candidate.email]).send()


class CandidateUpdateAPIView(generics.RetrieveUpdateAPIView):
    """
       Update candidate information.

       Methods:
           perform_update(self, serializer): Perform the update of candidate information.

    """
    queryset = CandidateModel.objects.all()
    serializer_class = CandidateUpdateSerializer
    parser_classes = (parsers.MultiPartParser,)

    def perform_update(self, serializer):
        candidate = serializer.instance
        candidate_approval = serializer.validated_data.get('candidate_approval')
        serializer.instance.candidate_approval = candidate_approval

        skills = serializer.validated_data.get('skills')
        serializer.instance.skills = skills

        languages = serializer.validated_data.get('languages')
        serializer.instance.languages = languages

        about = serializer.validated_data.get('about')
        serializer.instance.about = about

        serializer.instance.save(update_fields=['candidate_approval', 'skills', 'languages', 'about'])

        exist_appointment = AppointmentModel.objects.filter(candidate_id=candidate.id).first()

        resume = serializer.validated_data.get('resume')

        requirement = Requirement.objects.all()
        settings = SettingsModel.objects.all().first()
        educations = candidate.educationmodel_set.all()
        experiences_count = candidate.experiencesmodel_set.count()
        current_date = timezone.now()

        if resume:
            jalali_update_date = jdatetime.fromgregorian(datetime=candidate.update_at)
            formatted_update_date = jalali_update_date.strftime('%Y_%m_%d_%H.%M')
            file_path = f"{candidate.job}/{candidate.name}_{formatted_update_date}.pdf"
            candidate.resume.save(file_path, resume, save=True)

        if exist_appointment is None:
            if candidate.candidate_approval:
                candidate.score = calculate_skill_score(candidate, educations, requirement, experiences_count)
                candidate.save(update_fields=['score'])

                if candidate.score >= settings.pass_score:
                    try:
                        status_model = StatusModel.objects.get(candidate_id=candidate.id)
                        status_model.status = 'WI'
                        status_model.save()
                    except StatusModel.DoesNotExist:
                        StatusModel.objects.create(candidate_id=candidate.id, status='WI')

                    schedule_interviews(candidate, settings.interview_duration_hours, settings.start_work_time,
                                        settings.end_work_time, current_date)
                elif candidate.score is None or candidate.score <= settings.pass_score:
                    try:
                        status_model = StatusModel.objects.get(candidate_id=candidate.id)
                        status_model.status = 'R'
                        status_model.save()
                    except StatusModel.DoesNotExist:
                        StatusModel.objects.create(candidate_id=candidate.id, status='R')

                    print(f'Your resume rejected', 'Hello, we may reach out to you again in the future')

                    # EmailMessage(f'Your resume rejected', 'Hello, we may reach out to you again in the future',
                    #              config.EMAIL_HOST_USER, [candidate.email]).send()

        elif candidate.candidate_approval == False:
            try:
                status_model = StatusModel.objects.get(candidate_id=candidate.id)
                status_model.status = 'R'
                status_model.save()
            except StatusModel.DoesNotExist:
                StatusModel.objects.create(candidate_id=candidate.id, status='R')

            appointment = AppointmentModel.objects.filter(candidate_id=candidate.id).first()
            if appointment:
                appointment.delete()


class AppointmentViewSet(viewsets.ModelViewSet):
    queryset = AppointmentModel.objects.all()
    serializer_class = AppointmentSerializer
    # permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        exist_appointment = AppointmentModel.objects.filter(candidate_id=request.data.get('candidate')).first()
        if exist_appointment:
            return Response(
                {'success': True, 'status': 204, 'message': 'An appointment have already created for this candidate'})
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=201, headers=headers)


class SettingsViewSet(viewsets.ModelViewSet):
    queryset = SettingsModel.objects.all()
    serializer_class = SettingsSerializer


class PDFScoreAPIView(viewsets.ModelViewSet):
    queryset = CandidateModel.objects.all().order_by('-score')
    serializer_class = PDFScoreSerializer